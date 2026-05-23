"""Export agent — builds Excel workbook, uploads to GCS, returns signed URL."""
import io
import json
import logging
from datetime import datetime, timezone, timedelta
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from google.cloud import storage

from config import get_settings

settings = get_settings()
logger = logging.getLogger(__name__)

# Column definitions: (header, job_dict_key, width)
JOB_COLUMNS = [
    ("Title",             "title",            30),
    ("Company",           "company",          22),
    ("Location",          "location",         22),
    ("Country",           "country",          14),
    ("Salary",            "salary",           18),
    ("Job Type",          "job_type",         14),
    ("Source",            "source",           12),
    ("Posted Date",       "posted_date",      14),
    ("URL",               "url",              40),
    ("JD Full Text",      "jd_full_text",     60),
    ("Score",             "score",            10),
    ("Match Summary",     "match_summary",    45),
    ("Missing Keywords",  "missing_keywords", 30),
    ("Skill Gaps",        "skill_gaps",       40),
    ("Apply Priority",    "apply_priority",   14),
    ("Run ID",            "run_id",           20),
    ("User ID",           "user_id",          20),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

PRIORITY_FILLS = {
    "High":   PatternFill("solid", fgColor="C6EFCE"),
    "Medium": PatternFill("solid", fgColor="FFEB9C"),
    "Low":    PatternFill("solid", fgColor="FFC7CE"),
}


def _score_fill(score: int) -> Optional[PatternFill]:
    if score >= 80:
        return PatternFill("solid", fgColor="C6EFCE")
    elif score >= 60:
        return PatternFill("solid", fgColor="FFEB9C")
    elif score >= 40:
        return PatternFill("solid", fgColor="FFF2CC")
    return PatternFill("solid", fgColor="FFC7CE")


def build_excel(
    jobs: list[dict],
    run_meta: dict,
) -> bytes:
    """Build an Excel workbook in memory and return bytes."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Jobs ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Jobs"
    ws.freeze_panes = "A2"

    # Header row
    for col_idx, (header, _, width) in enumerate(JOB_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, job in enumerate(jobs, start=2):
        for col_idx, (_, key, _) in enumerate(JOB_COLUMNS, start=1):
            value = job.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Colour score column (col 11)
            if key == "score" and isinstance(value, int):
                cell.fill = _score_fill(value)

            # Colour apply_priority column (col 15)
            if key == "apply_priority" and value in PRIORITY_FILLS:
                cell.fill = PRIORITY_FILLS[value]

            # Make URL a hyperlink
            if key == "url" and value:
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")

    ws.auto_filter.ref = f"A1:{get_column_letter(len(JOB_COLUMNS))}1"

    # ── Sheet 2: Run Summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Run Summary")
    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 50

    summary_rows = [
        ("Run ID",              run_meta.get("run_id", "")),
        ("User ID",             run_meta.get("user_id", "")),
        ("Timestamp",           run_meta.get("timestamp", "")),
        ("Keywords",            ", ".join(run_meta.get("keywords", []))),
        ("Countries",           ", ".join(run_meta.get("countries", []))),
        ("APIs Used",           ", ".join(run_meta.get("apis_used", []))),
        ("Model Used",          run_meta.get("model_used", "")),
        ("Total Raw Jobs",      run_meta.get("total_raw_jobs", 0)),
        ("After Dedup",         run_meta.get("after_dedup", 0)),
        ("After Score Filter",  run_meta.get("after_score_filter", 0)),
        ("Min Score Threshold", run_meta.get("min_score", 40)),
        ("Max Results / API",   run_meta.get("max_results", 20)),
        ("Job Types",           ", ".join(run_meta.get("job_types", []) or [])),
        ("Work Modes",          ", ".join(run_meta.get("work_modes", []) or [])),
        ("Posted Within Days",  run_meta.get("posted_within_days", 7)),
        ("Est. Token Cost USD", run_meta.get("est_token_cost", "~$0.003")),
    ]

    for r, (label, value) in enumerate(summary_rows, start=1):
        a = ws2.cell(row=r, column=1, value=label)
        a.font = Font(bold=True)
        ws2.cell(row=r, column=2, value=value)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def upload_to_gcs(
    excel_bytes: bytes,
    run_meta: dict,
) -> tuple[str, str]:
    """
    Upload Excel to GCS.
    Returns (gcs_path, signed_url).
    """
    client = storage.Client()
    bucket = client.bucket(settings.GCS_BUCKET_NAME)

    user_id = run_meta["user_id"]
    run_id = run_meta["run_id"]
    date_str = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d")

    blob_name = f"users/{user_id}/{date_str}/run_{run_id}.xlsx"
    blob = bucket.blob(blob_name)
    blob.upload_from_string(excel_bytes, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    # Also save run meta JSON
    meta_blob = bucket.blob(f"users/{user_id}/{date_str}/run_{run_id}_meta.json")
    meta_blob.upload_from_string(json.dumps(run_meta, default=str), content_type="application/json")

    gcs_path = f"gs://{settings.GCS_BUCKET_NAME}/{blob_name}"

    # Signed URL valid for 15 min
    expiry = timedelta(minutes=settings.SIGNED_URL_EXPIRY_MINUTES)
    signed_url = blob.generate_signed_url(expiration=expiry, method="GET", version="v4")

    logger.info("Exported to %s", gcs_path)
    return gcs_path, signed_url
